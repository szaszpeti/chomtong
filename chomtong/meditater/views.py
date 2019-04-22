from django.shortcuts import render
from django.views.generic import (TemplateView,
                                  ListView,
                                  DetailView,
                                  DeleteView,
                                  CreateView,
                                  UpdateView)
from django.shortcuts import render, get_object_or_404
from .models import Meditater

from django.urls import reverse_lazy, reverse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q
from .forms import UploadFileForm
from django.core.mail import send_mail
import django_excel as excel
from smtplib import SMTP_SSL as SMTP
from email.mime.text import MIMEText

import openpyxl
from django.http import HttpResponse

import pandas as pd

from datetime import datetime
from datetime import timedelta








# Create your views here.
class HomeView(LoginRequiredMixin, TemplateView):
    model = Meditater
    template_name = "meditater/index.html"

    def meditaters(self):
        return Meditater.objects.all()
    #!!!!USE "VIEW.MEDITATERS" on the template


class FindView(LoginRequiredMixin, TemplateView):
    template_name = "meditater/find.html"



class MeditaterListView(LoginRequiredMixin, ListView):
    model = Meditater
    ordering = ["state"]
    template_name = "meditater/meditater_list.html"
    #context_object_name = 'post'#


class MeditaterDetailView(LoginRequiredMixin, DetailView):
    model = Meditater
    template_name = "meditater/meditater_detail.html"


class MeditaterDeleteView(LoginRequiredMixin, DeleteView):
    model = Meditater
    template_name = "meditater/meditater_delete.html"
    success_url = reverse_lazy('meditater:meditater_list')


class MeditaterUpdateView(LoginRequiredMixin, UpdateView):
    model = Meditater
    template_name = "meditater/meditater_edit.html"
    fields = ['course_type',
              'name',
              'state',
              'email',
              'gender',
              'born',
              'profession']
    success_url = reverse_lazy('meditater:meditater_list')

class MeditaterCreateView(LoginRequiredMixin, CreateView):
    model = Meditater
    template_name = "meditater/meditater_create.html"
    fields = ['course_type',
              'name',
              'state',
              'email',
              'gender',
              'born',
              'profession']
    success_url = reverse_lazy('meditater:meditater_list')

def find_index(request):
    print ("THis is the request from find", request)
    queryset_list = Meditater.objects.all().order_by("state")


    query = request.GET.get("q")
    print (query)
    if query:
        queryset_list = queryset_list.filter(
            Q(born__icontains=query) |
            Q(name__icontains=query) |
            Q(state__icontains=query) |
            Q(email__icontains=query) |
            Q(profession__icontains=query)

        ).distinct()
        number_found = len(queryset_list)

        return render(request, 'meditater/query_result.html', {'queryset': queryset_list, "query":query, "number":number_found})




def email_view(request, pk):
    meditater = get_object_or_404(Meditater, id=pk)

    if request.method == 'POST':
        toaddr = meditater.email

        subject = request.POST.get('tema')
        message = request.POST.get('uzenet')
        print(subject, message, toaddr)

        fromaddr = "szaszpeti01@gmail.com"
        msg = MIMEText(message, 'plain')
        msg['To'] = toaddr
        msg['Subject'] = subject

        server = SMTP('smtp.gmail.com')
        server.login(fromaddr, "Jimmypage")
        server.sendmail(fromaddr, toaddr, msg.as_string())
        server.quit()

    return render(request, "meditater/email_formating.html", {'name':meditater.name,
                                                              'email':meditater.email,
                                                              'country':meditater.state})





def export_to_excel_from_search(request):
    queryset_list = Meditater.objects.all()

    query = request.GET.get("q")
    print (query)
    if query:
        queryset_list = queryset_list.filter(
            Q(born__icontains=query) |
            Q(name__icontains=query) |
            Q(state__icontains=query) |
            Q(email__icontains=query) |
            Q(profession__icontains=query)

        ).distinct()

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-meditaters.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = openpyxl.Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Meditaters'

        # Define the titles for columns
        columns = [
            'year',
            'course_type',
            'name',
            'state',
            'email',
            'gender',
            'born',
            'profession',
        ]
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all movies
        for meditater in queryset_list:
            row_num += 1

            # Define the data for each cell in the row
            row = [
                meditater.year,
                meditater.course_type,
                meditater.name,
                meditater.state,
                meditater.email,
                meditater.gender,
                meditater.born,
                meditater.profession,
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        print("Excel file saved..........")
        print(response)
        return response

def meditater_statistic(request):
    pass
    df = pd.DataFrame(list(Meditater.objects.all().values()))
    state_dict=df['state'].value_counts().to_dict()
    profession_dict = df['profession'].value_counts().to_dict()
    gender_dict = df['gender'].value_counts().to_dict()
    born_dict = df['born'].value_counts().to_dict()
    year_dict = df['year'].value_counts().to_dict()



    return render(request, "meditater/statistic.html", {'state_dict':state_dict,
                                                        'profession_dict':profession_dict,
                                                        'gender_dict': gender_dict,
                                                        'born_dict': born_dict,
                                                        'year_dict': year_dict})
def statistic_country(request):
    pass
    df = pd.DataFrame(list(Meditater.objects.all().values()))
    state_dict=df['state'].value_counts().to_dict()

    return render(request, "meditater/statistic.html", {'state_dict':state_dict})


def statistic_profession(request):
    pass
    df = pd.DataFrame(list(Meditater.objects.all().values()))
    profession_dict = df['profession'].value_counts().to_dict()

    return render(request, "meditater/statistic.html", {
                                                        'profession_dict':profession_dict,
                                                       })
def statistic_gender(request):
    pass
    df = pd.DataFrame(list(Meditater.objects.all().values()))
    gender_dict = df['gender'].value_counts().to_dict()


    return render(request, "meditater/statistic.html", {
                                                        'gender_dict': gender_dict})


def statistic_born(request):
    pass
    df = pd.DataFrame(list(Meditater.objects.all().values()))
    born_dict = df['born'].value_counts().to_dict()

    return render(request, "meditater/statistic.html", {
                                                        'born_dict': born_dict})



def statistic_year(request):
    pass
    df = pd.DataFrame(list(Meditater.objects.all().values()))
    year_dict = df['year'].value_counts().to_dict()



    return render(request, "meditater/statistic.html", {'year_dict': year_dict})


def upload_file(request):
    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            filehandle = request.FILES['file']
            return excel.make_response(filehandle.get_sheet(), "csv",
                                       file_name="download")
    else:
        form = UploadFileForm()
    return render(
        request,
        'meditater/upload_form.html',
        {
            'form': form,
            'title': 'Excel file upload and download example',
            'header': ('Please choose any excel file ' +
                       'from your cloned repository:')
        })

